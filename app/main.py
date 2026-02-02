import json
import os
import re
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

import httpx
from fastapi import Body, FastAPI, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

ALLOWED_CATEGORIES = ["전략", "데이터", "AI", "자동화", "운영", "보안", "성과"]

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_PATH = BASE_DIR / "data" / "glossary.json"
DRAFTS_PATH = BASE_DIR / "data" / "drafts.json"


def load_env_file():
    env_path = BASE_DIR / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip())


load_env_file()

LLM_MODE = os.environ.get("LLM_MODE", "off").strip()  # off|azure_openai|hchat
LLM_ENDPOINT = os.environ.get("LLM_ENDPOINT", "").strip()
LLM_API_KEY = os.environ.get("LLM_API_KEY", "").strip()
LLM_API_KEY_HEADER = os.environ.get("LLM_API_KEY_HEADER", "api-key").strip()
LLM_DEPLOYMENT = os.environ.get("LLM_DEPLOYMENT", "").strip()
LLM_API_VERSION = os.environ.get("LLM_API_VERSION", "2024-02-15-preview").strip()

app = FastAPI(title="Glossary WebApp")
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")

tpl_env = Environment(
    loader=FileSystemLoader(str(BASE_DIR / "templates")),
    autoescape=select_autoescape(["html", "xml"]),
)


def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def load_glossary() -> List[Dict[str, Any]]:
    if not DATA_PATH.exists():
        return []
    return json.loads(DATA_PATH.read_text(encoding="utf-8"))


def save_glossary(items: List[Dict[str, Any]]):
    DATA_PATH.write_text(json.dumps(items, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def load_drafts() -> List[Dict[str, Any]]:
    # Deprecated: drafts are now saved directly into glossary.json (no admin approval).
    if not DRAFTS_PATH.exists():
        return []
    return json.loads(DRAFTS_PATH.read_text(encoding="utf-8"))


def save_drafts(items: List[Dict[str, Any]]):
    # Deprecated
    DRAFTS_PATH.write_text(json.dumps(items, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def find_term(term: str) -> Optional[Dict[str, Any]]:
    t = _norm(term)
    for item in load_glossary():
        if _norm(item.get("kr", "")) == t or _norm(item.get("en", "")) == t:
            return {**item, "source": item.get("createdBy", "glossary")}
    return None


def search_terms(q: str, category: str = "") -> List[Dict[str, Any]]:
    qn = _norm(q)
    catn = _norm(category)
    out: List[Dict[str, Any]] = []

    if not qn and not catn:
        return out

    # Search across glossary (all items are "confirmed" by default).
    items = load_glossary()

    for item in items:
        if catn and _norm(item.get("category", "")) != catn:
            continue

        if qn:
            hay = " ".join([
                item.get("kr", ""),
                item.get("en", ""),
                item.get("category", ""),
                item.get("oneLine", ""),
                " ".join(item.get("kpi", []) or []),
            ])
            if qn not in _norm(hay):
                continue

        out.append({
            "kr": item.get("kr"),
            "en": item.get("en"),
            "category": item.get("category"),
            "oneLine": item.get("oneLine"),
            "source": item.get("createdBy", "glossary"),
        })

    return out[:200]


def build_prompt(term: str) -> str:
    return (
        "당신은 제조(스마트팩토리/품질/설비) 맥락의 AI/DX 용어집 작성자입니다.\n"
        "아래 용어에 대해 '초안'을 작성하세요.\n\n"
        f"용어: {term}\n\n"
        "출력은 반드시 JSON만 반환하세요. 추가 텍스트 금지.\n"
        "스키마:\n"
        "{\n"
        "  \"kr\": \"\",\n"
        "  \"en\": \"\",\n"
        "  \"category\": \"전략|데이터|AI|자동화|운영|보안|성과\",\n"
        "  \"oneLine\": \"1~2문장\",\n"
        "  \"example\": \"제조 현장 예시 1개\",\n"
        "  \"kpi\": [\"OEE\",\"불량률\",\"리드타임\",\"OTD\",\"원가\",\"에너지\"],\n"
        "  \"confusions\": [\"유사 용어 1~3개\"],\n"
        "  \"ask\": [\"임원이 물어볼 질문 2개\"]\n"
        "}\n"
    )


def _split_list(s: Any) -> List[str]:
    if s is None:
        return []
    if isinstance(s, list):
        return [str(x).strip() for x in s if str(x).strip()]
    s2 = str(s).strip()
    if not s2:
        return []
    # allow comma or newline separated
    parts = re.split(r"[\n,]+", s2)
    return [p.strip() for p in parts if p.strip()]


def _merge_keep_existing(dst: Dict[str, Any], src: Dict[str, Any]):
    """Fill only missing/empty fields in dst from src."""
    for k, v in src.items():
        if v is None:
            continue
        if isinstance(v, str):
            if not v.strip():
                continue
        if isinstance(v, list):
            if len(v) == 0:
                continue

        cur = dst.get(k)
        if cur is None:
            dst[k] = v
            continue
        if isinstance(cur, str) and not cur.strip():
            dst[k] = v
            continue
        if isinstance(cur, list) and len(cur) == 0:
            dst[k] = v
            continue


def llm_generate(term: str) -> Dict[str, Any]:
    if LLM_MODE == "off":
        raise RuntimeError("LLM is disabled (LLM_MODE=off)")
    if not LLM_ENDPOINT or not LLM_API_KEY:
        raise RuntimeError("Missing LLM_ENDPOINT or LLM_API_KEY")

    prompt = build_prompt(term)

    # We support two generic modes:
    # - azure_openai: endpoint like https://{resource}.openai.azure.com
    # - hchat: custom endpoint that accepts OpenAI-compatible /chat/completions (best effort)

    headers = {LLM_API_KEY_HEADER: LLM_API_KEY}

    if LLM_MODE == "azure_openai":
        if not LLM_DEPLOYMENT:
            raise RuntimeError("Missing LLM_DEPLOYMENT for azure_openai")
        url = f"{LLM_ENDPOINT.rstrip('/')}/openai/deployments/{LLM_DEPLOYMENT}/chat/completions?api-version={LLM_API_VERSION}"
    else:
        # assume endpoint is full URL to /chat/completions
        url = LLM_ENDPOINT

    body = {
        "messages": [
            {"role": "system", "content": "Return only JSON."},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.2,
    }

    with httpx.Client(timeout=30.0) as client:
        r = client.post(url, headers=headers, json=body)
        r.raise_for_status()
        data = r.json()

    # best effort parsing
    content = None
    try:
        content = data["choices"][0]["message"]["content"]
    except Exception:
        raise RuntimeError(f"Unexpected LLM response shape: {str(data)[:500]}")

    try:
        obj = json.loads(content)
    except Exception:
        raise RuntimeError(f"LLM did not return valid JSON: {content[:500]}")

    # minimal normalization
    obj.setdefault("kr", term)
    obj.setdefault("category", "AI")
    obj.setdefault("oneLine", "")
    obj.setdefault("example", "")
    obj.setdefault("kpi", [])
    obj.setdefault("confusions", [])
    obj.setdefault("ask", [])
    return obj


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    tpl = tpl_env.get_template("index.html")
    return tpl.render(llm_enabled=(LLM_MODE != "off"))


@app.get("/api/categories")
def api_categories():
    # fixed list for UI consistency, but include any extra categories found in data
    cats = set(ALLOWED_CATEGORIES)
    for it in load_glossary():
        c = (it.get("category") or "").strip()
        if c:
            cats.add(c)
    ordered = [c for c in ALLOWED_CATEGORIES if c in cats]
    extras = sorted([c for c in cats if c not in ordered])
    return {"categories": ordered + extras}


@app.get("/api/search")
def api_search(q: str = "", category: str = ""):
    return {"results": search_terms(q, category=category)}


@app.get("/api/term")
def api_term(term: str = ""):
    item = find_term(term)
    if not item:
        return JSONResponse(status_code=404, content={"error": "NOT_FOUND"})
    return item


@app.get("/api/export.xlsx")
def api_export_xlsx():
    """Export the current glossary (and drafts) to an .xlsx file.

    This is intentionally server-side so end users can download via GUI.
    """

    items = load_glossary()

    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"

    headers = [
        "용어(KR)",
        "약어/EN",
        "분류",
        "한줄 정의",
        "예시",
        "KPI",
        "혼동되는 용어",
        "현업 질문(체크리스트)",
        "출처",
    ]
    ws.append(headers)

    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for it in items:
        src = it.get("createdBy", "USER")
        ws.append(
            [
                it.get("kr", ""),
                it.get("en", ""),
                it.get("category", ""),
                it.get("oneLine", ""),
                it.get("example", ""),
                ", ".join(it.get("kpi") or []),
                ", ".join(it.get("confusions") or []),
                "\n".join(it.get("ask") or []),
                src,
            ]
        )

    # Simple readability formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # column widths (rough)
    widths = [22, 22, 12, 70, 60, 25, 30, 55, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "AI_DX_Glossary_Manufacturing.xlsx"
    headers_out = {"Content-Disposition": f"attachment; filename=\"{filename}\""}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_out,
    )


@app.post("/api/upload.xlsx")
async def api_upload_xlsx(file: UploadFile = File(...), fillMissing: str = Form("on")):
    """Bulk upload terms from an Excel file.

    - Missing columns can be auto-filled by LLM (if enabled).
    - Uploaded rows are merged into glossary.json (the "OK" dataset).

    Expected headers (row 1) - either Korean or keys:
      용어(KR)|kr, 약어/EN|en, 분류|category, 한줄 정의|oneLine, 예시|example,
      KPI|kpi, 혼동되는 용어|confusions, 현업 질문(체크리스트)|ask
    """

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return JSONResponse(status_code=400, content={"error": "XLSX_ONLY"})

    content = await file.read()
    wb = load_workbook(filename=BytesIO(content))
    ws = wb.active

    # Build header map
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_row = [str(x).strip() if x is not None else "" for x in header_row]

    col_map: Dict[str, int] = {}
    aliases = {
        "kr": ["용어(KR)", "용어", "KR", "kr"],
        "en": ["약어/EN", "EN", "en", "약어"],
        "category": ["분류", "category"],
        "oneLine": ["한줄 정의", "정의", "oneLine"],
        "example": ["예시", "example"],
        "kpi": ["KPI", "kpi"],
        "confusions": ["혼동되는 용어", "confusions"],
        "ask": ["현업 질문(체크리스트)", "임원 질문", "ask"],
    }

    for i, h in enumerate(header_row):
        hn = h.strip()
        if not hn:
            continue
        for key, names in aliases.items():
            if hn in names:
                col_map[key] = i

    if "kr" not in col_map:
        return JSONResponse(status_code=400, content={"error": "MISSING_KR_COLUMN", "headers": header_row})

    glossary = load_glossary()

    def find_index(term: str) -> Optional[int]:
        tn = _norm(term)
        for idx, it in enumerate(glossary):
            if _norm(it.get("kr", "")) == tn or (_norm(it.get("en", "")) and _norm(it.get("en", "")) == tn):
                return idx
        return None

    fill = fillMissing.lower() in ("1", "true", "on", "yes")

    report = {"added": 0, "updated": 0, "filledByLLM": 0, "skipped": 0, "errors": []}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            kr = row[col_map["kr"]].value if "kr" in col_map else ""
            kr = str(kr).strip() if kr is not None else ""
            if not kr:
                report["skipped"] += 1
                continue

            entry: Dict[str, Any] = {
                "kr": kr,
                "en": str(row[col_map["en"]].value).strip() if ("en" in col_map and row[col_map["en"]].value is not None) else "",
                "category": str(row[col_map["category"]].value).strip() if ("category" in col_map and row[col_map["category"]].value is not None) else "",
                "oneLine": str(row[col_map["oneLine"]].value).strip() if ("oneLine" in col_map and row[col_map["oneLine"]].value is not None) else "",
                "example": str(row[col_map["example"]].value).strip() if ("example" in col_map and row[col_map["example"]].value is not None) else "",
                "kpi": _split_list(row[col_map["kpi"]].value) if "kpi" in col_map else [],
                "confusions": _split_list(row[col_map["confusions"]].value) if "confusions" in col_map else [],
                "ask": _split_list(row[col_map["ask"]].value) if "ask" in col_map else [],
            }

            # normalize category if empty
            if not entry.get("category"):
                entry["category"] = "AI"

            # Auto-fill missing parts by LLM (never overwrite user-provided values)
            if fill and (LLM_MODE != "off"):
                needs = any([
                    not entry.get("en"),
                    not entry.get("category"),
                    not entry.get("oneLine"),
                    not entry.get("example"),
                    not entry.get("kpi"),
                    not entry.get("confusions"),
                    not entry.get("ask"),
                ])
                if needs:
                    gen = llm_generate(kr)
                    # ensure list shapes
                    gen["kpi"] = _split_list(gen.get("kpi"))
                    gen["confusions"] = _split_list(gen.get("confusions"))
                    gen["ask"] = _split_list(gen.get("ask"))
                    _merge_keep_existing(entry, gen)
                    report["filledByLLM"] += 1

            idx = find_index(kr)
            if idx is None and entry.get("en"):
                idx = find_index(entry.get("en", ""))

            if idx is None:
                glossary.append(entry)
                report["added"] += 1
            else:
                # merge: keep existing values, fill missing from uploaded/LLM
                existing = glossary[idx]
                merged = dict(existing)
                _merge_keep_existing(merged, entry)
                glossary[idx] = merged
                report["updated"] += 1

        except Exception as e:
            report["errors"].append({"row": row_idx, "error": str(e)})

    save_glossary(glossary)
    return {"ok": True, "report": report, "count": len(glossary)}


@app.post("/api/save")
def api_save(payload: Dict[str, Any] = Body(...)):
    """Upsert a glossary item. User can edit freely (no approval workflow)."""
    kr = (payload.get("kr") or "").strip()
    if not kr:
        return JSONResponse(status_code=400, content={"error": "MISSING_KR"})

    item: Dict[str, Any] = {
        "kr": kr,
        "en": (payload.get("en") or "").strip(),
        "category": (payload.get("category") or "").strip(),
        "oneLine": (payload.get("oneLine") or "").strip(),
        "example": (payload.get("example") or "").strip(),
        "kpi": _split_list(payload.get("kpi")),
        "confusions": _split_list(payload.get("confusions")),
        "ask": _split_list(payload.get("ask")),
        "createdBy": (payload.get("createdBy") or "USER").strip() or "USER",
    }
    if not item["category"]:
        item["category"] = "AI"

    glossary = load_glossary()
    idx = None
    tn = _norm(kr)
    for i, it in enumerate(glossary):
        if _norm(it.get("kr", "")) == tn:
            idx = i
            break

    if idx is None:
        glossary.append(item)
    else:
        glossary[idx] = item

    save_glossary(glossary)
    return {"ok": True, "item": find_term(kr)}


@app.post("/api/draft")
def api_draft(term: str = Form(...)):
    """Generate a new term by LLM and save it immediately as confirmed."""
    term = term.strip()
    if not term:
        return JSONResponse(status_code=400, content={"error": "EMPTY"})

    existing = find_term(term)
    if existing:
        return {"alreadyExists": True, "item": existing}

    obj = llm_generate(term)
    obj["createdBy"] = "LLM"

    glossary = load_glossary()
    glossary.append(obj)
    save_glossary(glossary)

    return {"ok": True, "item": find_term(obj.get("kr") or term)}
