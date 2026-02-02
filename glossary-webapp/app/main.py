import json
import os
import re
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

import httpx
from fastapi import FastAPI, Form, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_PATH = BASE_DIR / "data" / "glossary.json"
DRAFTS_PATH = BASE_DIR / "data" / "drafts.json"


def load_env_file():
    env_path = BASE_DIR / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip())


load_env_file()

LLM_MODE = os.environ.get("LLM_MODE", "off").strip()  # off|azure_openai|hchat
LLM_ENDPOINT = os.environ.get("LLM_ENDPOINT", "").strip()
LLM_API_KEY = os.environ.get("LLM_API_KEY", "").strip()
LLM_API_KEY_HEADER = os.environ.get("LLM_API_KEY_HEADER", "api-key").strip()
LLM_DEPLOYMENT = os.environ.get("LLM_DEPLOYMENT", "").strip()
LLM_API_VERSION = os.environ.get("LLM_API_VERSION", "2024-02-15-preview").strip()

app = FastAPI(title="Glossary WebApp")
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")

tpl_env = Environment(
    loader=FileSystemLoader(str(BASE_DIR / "templates")),
    autoescape=select_autoescape(["html", "xml"]),
)


def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def load_glossary() -> List[Dict[str, Any]]:
    if not DATA_PATH.exists():
        return []
    return json.loads(DATA_PATH.read_text(encoding="utf-8"))


def load_drafts() -> List[Dict[str, Any]]:
    if not DRAFTS_PATH.exists():
        return []
    return json.loads(DRAFTS_PATH.read_text(encoding="utf-8"))


def save_drafts(items: List[Dict[str, Any]]):
    DRAFTS_PATH.write_text(json.dumps(items, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def find_term(term: str) -> Optional[Dict[str, Any]]:
    t = _norm(term)
    for item in load_glossary():
        if _norm(item.get("kr", "")) == t or _norm(item.get("en", "")) == t:
            return {**item, "source": "glossary"}
    for item in load_drafts():
        if _norm(item.get("kr", "")) == t or _norm(item.get("en", "")) == t:
            return {**item, "source": "draft"}
    return None


def search_terms(q: str) -> List[Dict[str, Any]]:
    qn = _norm(q)
    out = []
    if not qn:
        return out
    for item in load_glossary():
        hay = " ".join([
            item.get("kr", ""),
            item.get("en", ""),
            item.get("category", ""),
            item.get("oneLine", ""),
            " ".join(item.get("kpi", []) or []),
        ])
        if qn in _norm(hay):
            out.append({"kr": item.get("kr"), "en": item.get("en"), "category": item.get("category"), "oneLine": item.get("oneLine")})
    return out[:50]


def build_prompt(term: str) -> str:
    return (
        "당신은 제조(스마트팩토리/품질/설비) 맥락의 AI/DX 용어집 작성자입니다.\n"
        "아래 용어에 대해 '초안'을 작성하세요.\n\n"
        f"용어: {term}\n\n"
        "출력은 반드시 JSON만 반환하세요. 추가 텍스트 금지.\n"
        "스키마:\n"
        "{\n"
        "  \"kr\": \"\",\n"
        "  \"en\": \"\",\n"
        "  \"category\": \"전략|데이터|AI|자동화|운영|보안|성과\",\n"
        "  \"oneLine\": \"1~2문장\",\n"
        "  \"example\": \"제조 현장 예시 1개\",\n"
        "  \"kpi\": [\"OEE\",\"불량률\",\"리드타임\",\"OTD\",\"원가\",\"에너지\"],\n"
        "  \"confusions\": [\"유사 용어 1~3개\"],\n"
        "  \"ask\": [\"임원이 물어볼 질문 2개\"]\n"
        "}\n"
    )


def llm_generate(term: str) -> Dict[str, Any]:
    if LLM_MODE == "off":
        raise RuntimeError("LLM is disabled (LLM_MODE=off)")
    if not LLM_ENDPOINT or not LLM_API_KEY:
        raise RuntimeError("Missing LLM_ENDPOINT or LLM_API_KEY")

    prompt = build_prompt(term)

    # We support two generic modes:
    # - azure_openai: endpoint like https://{resource}.openai.azure.com
    # - hchat: custom endpoint that accepts OpenAI-compatible /chat/completions (best effort)

    headers = {LLM_API_KEY_HEADER: LLM_API_KEY}

    if LLM_MODE == "azure_openai":
        if not LLM_DEPLOYMENT:
            raise RuntimeError("Missing LLM_DEPLOYMENT for azure_openai")
        url = f"{LLM_ENDPOINT.rstrip('/')}/openai/deployments/{LLM_DEPLOYMENT}/chat/completions?api-version={LLM_API_VERSION}"
    else:
        # assume endpoint is full URL to /chat/completions
        url = LLM_ENDPOINT

    body = {
        "messages": [
            {"role": "system", "content": "Return only JSON."},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.2,
    }

    with httpx.Client(timeout=30.0) as client:
        r = client.post(url, headers=headers, json=body)
        r.raise_for_status()
        data = r.json()

    # best effort parsing
    content = None
    try:
        content = data["choices"][0]["message"]["content"]
    except Exception:
        raise RuntimeError(f"Unexpected LLM response shape: {str(data)[:500]}")

    try:
        obj = json.loads(content)
    except Exception:
        raise RuntimeError(f"LLM did not return valid JSON: {content[:500]}")

    # minimal normalization
    obj.setdefault("kr", term)
    obj.setdefault("category", "AI")
    obj.setdefault("oneLine", "")
    obj.setdefault("example", "")
    obj.setdefault("kpi", [])
    obj.setdefault("confusions", [])
    obj.setdefault("ask", [])
    return obj


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    tpl = tpl_env.get_template("index.html")
    return tpl.render(llm_enabled=(LLM_MODE != "off"))


@app.get("/api/search")
def api_search(q: str = ""):
    return {"results": search_terms(q)}


@app.get("/api/term")
def api_term(term: str = ""):
    item = find_term(term)
    if not item:
        return JSONResponse(status_code=404, content={"error": "NOT_FOUND"})
    return item


@app.get("/api/export.xlsx")
def api_export_xlsx():
    """Export the current glossary (and drafts) to an .xlsx file.

    This is intentionally server-side so end users can download via GUI.
    """

    glossary = load_glossary()
    drafts = load_drafts()
    items = glossary + drafts

    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"

    headers = [
        "용어(KR)",
        "약어/EN",
        "분류",
        "한줄 정의",
        "예시",
        "KPI",
        "혼동되는 용어",
        "현업 질문(체크리스트)",
        "출처",
    ]
    ws.append(headers)

    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for it in items:
        src = "DRAFT" if it in drafts else "OK"
        ws.append(
            [
                it.get("kr", ""),
                it.get("en", ""),
                it.get("category", ""),
                it.get("oneLine", ""),
                it.get("example", ""),
                ", ".join(it.get("kpi") or []),
                ", ".join(it.get("confusions") or []),
                "\n".join(it.get("ask") or []),
                src,
            ]
        )

    # Simple readability formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # column widths (rough)
    widths = [22, 22, 12, 70, 60, 25, 30, 55, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "AI_DX_Glossary_Manufacturing.xlsx"
    headers_out = {"Content-Disposition": f"attachment; filename=\"{filename}\""}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_out,
    )


@app.post("/api/draft")
def api_draft(term: str = Form(...)):
    term = term.strip()
    if not term:
        return JSONResponse(status_code=400, content={"error": "EMPTY"})

    existing = find_term(term)
    if existing:
        return {"alreadyExists": True, "item": existing}

    obj = llm_generate(term)
    obj["status"] = "Draft"
    obj["createdBy"] = "LLM"

    drafts = load_drafts()
    drafts.insert(0, obj)
    save_drafts(drafts)

    return {"ok": True, "item": {**obj, "source": "draft"}}
